from datetime import datetime
import logging
import re

from bs4 import BeautifulSoup as BS
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.options import Options as FirefoxOptions

logging.basicConfig(
        level=logging.INFO,
        filename='parse_erz.log',
        filemode='w',
        format='%(asctime)s, %(levelname)s, %(name)s, %(message)s'
)

def main():
    wb = load_workbook('test.xlsx')
    sh = wb.active
    options = FirefoxOptions()
    options.add_argument("--headless")
    driver = webdriver.Firefox(options=options)
    driver.implicitly_wait(1)
    for i in range(794, 840): 
        url = sh.cell(row=i, column=5).value
        try:
            driver.get(url)
            try:
                for j in range(1, 5):
                    c = driver.find_element_by_css_selector(f'tab-header.ng-tns-c19-3:nth-child({j})')
                    c.click()
            except Exception as e:
                logging.info(f'{e}')
            html = driver.page_source
            soup = BS(html, 'html5lib')
            date_list = soup(string=re.compile('кв. 202'))
            years = []
            try:
                if len(date_list) > 1:
                    date = date_list[1].split()
                    k = 0
                    while len(date) == 3:
                        if k < len(date_list) - 1:
                            k = k + 1
                        else:
                            break
                        date = date_list[k].split()
                        if len(date) > 3:
                            break
                        if int(date[2]) >= datetime.now().year:
                            if date[0] == 'I':
                                date[0] = '1'
                            elif date[0] == 'II':
                                date[0] = '2'
                            elif date[0] == 'III':
                                date[0] = '3'
                            elif date[0] == 'IV':
                                date[0] = '4'
                            if date not in years:
                                years.append(date)
                    date_f = ''
                    for l in years:
                        date_f += (f'{l[0]}/{l[2]}|')
                    sh.cell(row=i, column=13).value = date_f[:-1]
            except Exception as e:
                logging.info(f'{e}')
            status = ''
            build = soup(string='Строится')
            passed = soup(string='Сдано')
            stopped = soup(string='Остановлено')
            if stopped != []:
                status = 'Остановлено'
            if passed == [] and stopped == []:
                status = 'Строится'
            elif passed != [] and years == []:
                status = 'Сдан'
            elif passed != [] and years != []:
                status = 'Строится, есть сданные'
            if passed == [] and build == [] and stopped == []:
                status = 'Пустая страница' 
            sh.cell(row=i, column=12).value = status
        except Exception as e:
            logging.info(f'{e}')
    wb.save('test.xlsx')
    driver.quit()


if __name__ == '__main__':
    main()