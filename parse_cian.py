import logging
import re
from datetime import datetime

import requests
from bs4 import BeautifulSoup as BS
from openpyxl import load_workbook

logging.basicConfig(
        level=logging.INFO,
        filename='parse_cian.log',
        filemode='w',
        format='%(asctime)s, %(levelname)s, %(name)s, %(message)s'
)


def main():
    wb = load_workbook('test.xlsx')
    sh = wb.active
    for i in range(2005, 4652):
        url = sh.cell(row=i, column=4).value
        date_f = ''
        status = 'Строится'
        try:
            response = requests.get(url)
            soup = BS(response.text, 'html5lib')
            date_list = soup(string=re.compile('квартал 20'))
            if date_list != []:
                date = date_list[0].split()
                if len(date) == 3:
                    n = 0
                    dates = []
                    statuses = []
                    while len(date) == 3:
                        date = date_list[n].split()
                        n += 1
                        if len(date) != 3:
                            break
                        statuses.append(date)
                        if date not in dates and int(date[2]) >= datetime.now().year:
                            dates.append(date)
                    for m in dates:
                        date_f += (f'{m[0]}/{m[2]}|')
                    a = 0
                    b = 0
                    for j in statuses:
                        if int(j[2]) < datetime.now().year or (j[0] == '1' and j[2] == '2021'):
                            b += 1
                        elif int(j[2]) >= datetime.now().year:
                            a += 1
                    try:
                        date_list_1 = soup(string=re.compile('есть сданные'), limit=1)
                        if a == 0 and b > 0:
                            status = 'Сдан'
                        elif (a > 0 and b > 0) or len(date_list_1[0].split()) < 10:
                            status = 'Строится, есть сданные'
                    except Exception as e:
                        logging.info(f'{e}', exc_info=True)
            try:
                c = soup.select('#newbuilding-card-desktop-frontend > main > section:nth-child(4) > div:nth-child(1) > div._02712f2b3b--leftColumn--18B6V > div._02712f2b3b--basicInfo--1YO3T > section > div._02712f2b3b--basic--18hNX > div > div > div > div > span')[0].text
                if c == 'Сдан': 
                    status = 'Сдан'
                    date_f = ''
                elif ('Сдача' in c) and len(c.split()) == 3:
                    date_f = (f'{c.split()[2]}|')
                elif ('Сдача' in c) and ('есть сданные' not in c) and len(c.split()) == 5:
                    date_f = (f'{c.split()[2]}/{c.split()[4]}|')
                elif ('есть сданные' in c) and len(c.split()) == 5:
                    status = 'Строится, есть сданные'
                    date_f = (f'{c.split()[2]}') 
                elif ('есть сданные' in c) and len(c.split()) == 7:
                    status = 'Строится, есть сданные'
                    date_f = (f'{c.split()[2]}/{c.split()[4]}')
            except Exception as e:
                logging.info(f'{e}', exc_info=True)
        except Exception as e:
            logging.info(f'{e}', exc_info=True)
        sh.cell(row=i, column=10).value = status
        sh.cell(row=i, column=11).value = date_f[:-1]
    wb.save('test.xlsx')


if __name__ == '__main__':
    main()
